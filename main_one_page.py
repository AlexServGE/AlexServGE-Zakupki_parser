from bs4 import BeautifulSoup
import requests
import common_info
import supplier_results
import rpec
from openpyxl import Workbook
import re

wb = Workbook()
ws = wb.active
columns_xls = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
row_chief = 0

# Йопромид https://zakupki.gov.ru/epz/order/extendedsearch/results.html?searchString=%D0%B9%D0%BE%D0%BF%D1%80%D0%BE%D0%BC%D0%B8%D0%B4&morphology=on&search-filter=%D0%94%D0%B0%D1%82%D0%B5+%D1%80%D0%B0%D0%B7%D0%BC%D0%B5%D1%89%D0%B5%D0%BD%D0%B8%D1%8F&pageNumber=1&sortDirection=false&recordsPerPage=_50&showLotsInfoHidden=false&sortBy=UPDATE_DATE&fz44=on&pc=on&currencyIdGeneral=-1
# Йогексол https://zakupki.gov.ru/epz/order/extendedsearch/results.html?searchString=%D0%B9%D0%BE%D0%B3%D0%B5%D0%BA%D1%81%D0%BE%D0%BB&morphology=on&search-filter=%D0%94%D0%B0%D1%82%D0%B5+%D1%80%D0%B0%D0%B7%D0%BC%D0%B5%D1%89%D0%B5%D0%BD%D0%B8%D1%8F&pageNumber=1&sortDirection=false&recordsPerPage=_50&showLotsInfoHidden=false&sortBy=UPDATE_DATE&fz44=on&pc=on&currencyIdGeneral=-1
# Гадобутрол https://zakupki.gov.ru/epz/order/extendedsearch/results.html?searchString=%D0%B3%D0%B0%D0%B4%D0%BE%D0%B1%D1%83%D1%82%D1%80%D0%BE%D0%BB&morphology=on&search-filter=%D0%94%D0%B0%D1%82%D0%B5+%D1%80%D0%B0%D0%B7%D0%BC%D0%B5%D1%89%D0%B5%D0%BD%D0%B8%D1%8F&pageNumber=1&sortDirection=false&recordsPerPage=_50&showLotsInfoHidden=false&sortBy=UPDATE_DATE&fz44=on&pc=on&currencyIdGeneral=-1

auctions = ["0318300163422000319"]
all_pages_list = ['common-info', 'supplier-results']

for auction in auctions:
    for el in all_pages_list:
        user_agent = ('Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:50.0) '
                      'Gecko/20100101 Firefox/50.0')
        main_p = requests.get(f'https://zakupki.gov.ru/epz/order/notice/ea20/view/{el}.html?regNumber={auction}',
                              timeout=(20, 10), headers={'User-Agent': user_agent})
        main_p_dec = main_p.content.decode(encoding='utf-8')
        soup = BeautifulSoup(main_p_dec, 'html.parser')
        if el == 'common-info':
            results_1_col = [auction, main_p.status_code, common_info.center_name(main_p_dec),
                             common_info.product_name(soup), common_info.sku_volume(soup), common_info.sku_price(soup),
                             common_info.sku_value(soup), common_info.form_concentration(soup)]
            max_len = len(common_info.form_concentration(soup))
            for idx, el_2 in enumerate(results_1_col):
                col = columns_xls[idx]
                row = 1 + row_chief
                if type(el_2) == list:
                    for el_3 in el_2:
                        ws[f'{col}{row}'] = el_3
                        row += 1
                else:
                    for row in range(1 + row_chief, max_len + 1 + row_chief):
                        ws[f'{col}{row}'] = el_2
            results_1_row = [common_info.dates(main_p_dec)]
            next_free_col_idx = columns_xls.index(col) + 1
            for el_2 in results_1_row:
                if type(el_2) == type(list()):
                    for idx_el_2 in range(len(el_2)):
                        col = columns_xls[next_free_col_idx]
                        for row in range(1 + row_chief, max_len + 1 + row_chief):
                            ws[f'{col}{row}'] = el_2[idx_el_2]
                        next_free_col_idx = columns_xls.index(col) + 1
            common_soup = soup
            print(main_p.status_code)
            print(common_info.center_name(main_p_dec))
            print(common_info.dates(main_p_dec))
            print(common_info.product_name(soup))
            print(common_info.sku_volume(soup))
            print(common_info.sku_price(soup))
            print(common_info.sku_value(soup))
            print(common_info.form_concentration(soup))
        elif el == 'supplier-results':
            # print(main_p.status_code)
            if supplier_results.all_participants_id(soup) == [] and supplier_results.all_participants_winner(
                    soup) == [] and supplier_results.all_participants_endvalue(soup, common_soup) == []:
                print(
                    'По окончании срока подачи заявок не подано ни одной заявки на участие в закупке (п. 3 ч. 1 ст. 52 Закона № 44-ФЗ)')
                col = columns_xls[next_free_col_idx]
                for row in range(1 + row_chief, max_len + 1 + row_chief):
                    ws[
                        f'{col}{row}'] = 'По окончании срока подачи заявок не подано ни одной заявки на участие в закупке (п. 3 ч. 1 ст. 52 Закона № 44-ФЗ)'
                next_free_col_idx = columns_xls.index(col) + 1
                row_chief += max_len
            else:
                # print(supplier_results.all_participants_id(soup))
                # print(supplier_results.all_participants_winner(soup))
                participants_endvalue = supplier_results.all_participants_endvalue(soup, common_soup)
                winner = supplier_results.winner(soup)
                print(winner)  # юр лицо победителя
                if winner == 'No data about winner YET':
                    user_agent = ('Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:50.0) Gecko/20100101 Firefox/50.0')
                    secondary_p = requests.get(
                        f'https://zakupki.gov.ru/epz/order/notice/rpec/common-info.html?regNumber={auction}0001',
                        timeout=(20, 10), headers={'User-Agent': user_agent})
                    if secondary_p.status_code == 404:
                        print('Страницы со статусом контракта пока не существует')
                    else:
                        secondary_p_dec = secondary_p.content.decode(encoding='utf-8')
                        soup = BeautifulSoup(secondary_p_dec, 'html.parser')
                        print(rpec.winner_info(soup))
                        winner = rpec.winner_info(soup)
                col = columns_xls[next_free_col_idx]
                for row in range(1 + row_chief, max_len + 1 + row_chief):
                    ws[f'{col}{row}'] = winner
                next_free_col_idx = columns_xls.index(col) + 1
                print(participants_endvalue)  # предложения участников
                for el_2 in participants_endvalue:
                    col = columns_xls[next_free_col_idx]
                    for row in range(1 + row_chief, max_len + 1 + row_chief):
                        ws[f'{col}{row}'] = el_2
                    next_free_col_idx = columns_xls.index(col) + 1
                row_chief += max_len  # Переход к следующей закупке

wb.save(r'tenders.xlsx')

#
# with open('new_file.txt','w',encoding='utf-8') as f:
#     f.write()
